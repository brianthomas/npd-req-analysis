
import psycopg2

# Program to load Excel workbook into postgresql database
DesiredWorksheet = 'RQMTs'

SheetToDBColMap = {
        'Identity' : None, 
        'DocName'  : 'docname', 
        'Name'     : None, 
        'ReqClass' : 'reqclass', 
        'Requirement SubClass' : 'reqsubclass',  
        'REQUIREMENT' : 'content', 
        }; 

def opendb ():
    conn = None
    try:
        print('Connecting to the PostgreSQL database...')
        conn = psycopg2.connect("dbname=npdreq user=postgres")

    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
        if conn is not None:
            conn.close()
            print('Database connection closed.')
        raise error

    return conn

def loaddb (conn, data):
    print ("Loading the database...")

    cur = None
    try:
        # Open a cursor to perform database operations
        cur = conn.cursor()

        # Execute a command: this creates a new table
        for datum in data:
            print (str(datum))
            keys = datum.keys()
            colstr = ','.join(keys)
            vals = []
            for col in datum.keys():
                val = datum[col]
                if val:
                    #vals.append("'"+datum[col].replace("'", "\\'").replace('"', '\\"')+"'")
                    #vals.append("'"+datum[col].replace("'", "\\\\'")+"'")
                    # replace single quotes with nothing to avoid loading errors
                    vals.append("'"+datum[col].replace("'", "")+"'")
                else:
                    vals.append("''")
            valstr = ','.join(vals)
            sql = "INSERT INTO requirements ("+colstr+") VALUES ("+valstr+");" 

            #print (sql)
            cur.execute(sql)

        conn.commit()

    finally:
        if cur : cur.close()
        conn.close()

def parse (filename):

    # load the workbook from file
    from openpyxl import load_workbook
    wb = load_workbook(filename)

    # grab the worksheet we want
    if DesiredWorksheet not in wb:
        raise Exception("Can't load desired worksheet from Excel file, bailing")
 
    ws = wb[DesiredWorksheet]

    # parse the column names
    colnames = []
    for row in ws.iter_rows(min_row=1, max_col=6, max_row=1):
        for cell in row:
            #print(cell.value)
            mappedVal = SheetToDBColMap[cell.value]
            colnames.append(mappedVal)

    # print (str(colnames))

    data = [] 
    # parse rows to memory array
    for row in ws.iter_rows(min_row=2, max_col=6):
        datum = {}
        i = 0
        for cell in row:
            colname = colnames[i]
            if colname:
                datum[colnames[i]] = cell.value
            i = i + 1

        data.append(datum)

    return data


if __name__ == '__main__':
    import argparse

    # Use nargs to specify how many arguments an option should take.
    ap = argparse.ArgumentParser(description='XLSM requirements db loader')
    ap.add_argument('-f', '--file', type=str, help='Name of the Excel file which holds the data')

    # parse argv
    opts = ap.parse_args()

    if not opts.file:
        print ("the --file <file> parameter must be specified")
        ap.print_usage()
        exit()

    loaddb(opendb(), parse(opts.file))

    print ("Finished loading")





